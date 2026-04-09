#!/usr/bin/env python3
"""Excel 耦合矩阵读写工具

从 Excel 工作簿中读取耦合矩阵数据（含单元格批注），
构建内存图模型，并支持新耦合点的回写。

用法:
    python scripts/excel_coupling_tool.py read  <excel_path> [--output <yaml_path>]
    python scripts/excel_coupling_tool.py write <excel_path> --source <yaml_path>
    python scripts/excel_coupling_tool.py query <yaml_path> --feature <feature_name>
"""
import argparse
import json
import sys
import os
from pathlib import Path

# --- 尝试导入 openpyxl；不可用时回退到 zipfile + XML 解析 ---
try:
    import openpyxl
    from openpyxl.comments import Comment
    HAS_OPENPYXL = True
except ImportError:
    HAS_OPENPYXL = False


def _read_comments_via_openpyxl(wb_path: str) -> dict:
    """使用 openpyxl 读取 Excel 工作簿中的所有批注。"""
    wb = openpyxl.load_workbook(wb_path)
    result = {}
    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        sheet_comments = {}
        for row in ws.iter_rows():
            for cell in row:
                if cell.comment:
                    coord = cell.coordinate
                    sheet_comments[coord] = {
                        "text": cell.comment.text.strip(),
                        "author": cell.comment.author or "",
                        "cell_value": str(cell.value) if cell.value is not None else "",
                        "row": cell.row,
                        "column": cell.column,
                    }
        if sheet_comments:
            result[sheet_name] = sheet_comments
    wb.close()
    return result


def _read_comments_via_zipfile(wb_path: str) -> dict:
    """使用 zipfile + XML 解析读取 Excel 批注（openpyxl 不可用时的回退方案）。"""
    import zipfile
    import xml.etree.ElementTree as ET

    result = {}
    ns = {
        "main": "http://schemas.openxmlformats.org/spreadsheetml/2006/main",
        "r": "http://schemas.openxmlformats.org/officeDocument/2006/relationships",
    }

    with zipfile.ZipFile(wb_path, "r") as zf:
        # 读取工作簿获取 sheet 名称映射
        wb_xml = ET.fromstring(zf.read("xl/workbook.xml"))
        sheets_elem = wb_xml.find("main:sheets", ns)
        sheet_names = {}
        if sheets_elem is not None:
            for s in sheets_elem.findall("main:sheet", ns):
                rid = s.get("{%s}id" % ns["r"])
                sheet_names[rid] = s.get("name")

        # 读取关系文件获取 sheet 文件路径
        rels_xml = ET.fromstring(zf.read("xl/_rels/workbook.xml.rels"))
        sheet_files = {}
        for rel in rels_xml:
            rid = rel.get("Id")
            target = rel.get("Target")
            if rid in sheet_names and target:
                sheet_files[sheet_names[rid]] = target

        # 遍历每个 sheet 的 comments 文件
        file_list = zf.namelist()
        for name in file_list:
            if name.startswith("xl/comments") and name.endswith(".xml"):
                comments_xml = ET.fromstring(zf.read(name))
                # 尝试确定对应的 sheet 名称
                sheet_idx = name.replace("xl/comments", "").replace(".xml", "")

                comment_list_elem = comments_xml.find("main:commentList", ns)
                if comment_list_elem is None:
                    continue

                sheet_comments = {}
                for c in comment_list_elem.findall("main:comment", ns):
                    ref = c.get("ref", "")
                    author_id = c.get("authorId", "0")
                    text_elem = c.find(".//main:t", ns)
                    text = text_elem.text.strip() if text_elem is not None and text_elem.text else ""
                    if text:
                        sheet_comments[ref] = {
                            "text": text,
                            "author": author_id,
                            "cell_value": "",
                            "row": 0,
                            "column": 0,
                        }

                if sheet_comments:
                    # 用序号作为 key（无法精确映射 sheet 名称时）
                    result[f"Sheet{sheet_idx}"] = sheet_comments

    return result


def read_coupling_matrix(wb_path: str) -> dict:
    """读取耦合矩阵 Excel，返回结构化数据。"""
    if not os.path.exists(wb_path):
        print(f"错误：文件不存在 - {wb_path}", file=sys.stderr)
        sys.exit(1)

    print(f"读取 Excel：{wb_path}")
    if HAS_OPENPYXL:
        print("使用 openpyxl 解析")
        comments = _read_comments_via_openpyxl(wb_path)
    else:
        print("openpyxl 不可用，使用 zipfile+XML 回退解析")
        comments = _read_comments_via_zipfile(wb_path)

    # 构建图模型
    graph = {
        "source_file": wb_path,
        "parser": "openpyxl" if HAS_OPENPYXL else "zipfile+xml",
        "sheets": {},
        "nodes": [],
        "edges": [],
    }

    all_features = set()
    edges = []

    for sheet_name, sheet_comments in comments.items():
        graph["sheets"][sheet_name] = {
            "comment_count": len(sheet_comments),
            "comments": {},
        }
        for coord, cdata in sheet_comments.items():
            comment_text = cdata["text"]
            graph["sheets"][sheet_name]["comments"][coord] = {
                "text": comment_text,
                "author": cdata.get("author", ""),
                "cell_value": cdata.get("cell_value", ""),
            }

            # 尝试将批注解析为耦合点
            # 耦合矩阵的批注格式通常是自由文本描述
            if _is_coupling_comment(comment_text):
                edge = {
                    "source_sheet": sheet_name,
                    "cell": coord,
                    "description": comment_text,
                    "strength": _infer_strength(comment_text),
                    "origin": "matrix-baseline",
                    "confirmed": True,
                }
                edges.append(edge)

    graph["edges"] = edges
    total_comments = sum(s["comment_count"] for s in graph["sheets"].values())
    coupling_count = len(edges)
    print(f"总批注数：{total_comments}，识别为耦合点：{coupling_count}")

    return graph


def _is_coupling_comment(text: str) -> bool:
    """判断批注是否为耦合点描述（vs 审阅批注/格式说明）。"""
    # 排除常见的非耦合批注
    skip_patterns = ["审阅", "格式", "TODO", "待确认", "已删除", "模板说明"]
    for p in skip_patterns:
        if p in text:
            return False
    # 耦合点批注通常包含功能/特性相关的描述
    if len(text) < 5:
        return False
    return True


def _infer_strength(text: str) -> str:
    """从批注文本推断耦合强度。"""
    strong_keywords = ["强耦合", "必须", "依赖", "前置", "阻塞"]
    weak_keywords = ["弱耦合", "可选", "建议", "参考"]
    for kw in strong_keywords:
        if kw in text:
            return "strong"
    for kw in weak_keywords:
        if kw in text:
            return "weak"
    return "normal"


def write_coupling_to_excel(wb_path: str, new_edges: list):
    """将新耦合点写回 Excel 矩阵（以批注形式）。"""
    if not HAS_OPENPYXL:
        print("错误：回写 Excel 需要 openpyxl 库", file=sys.stderr)
        print("请执行：pip install openpyxl", file=sys.stderr)
        sys.exit(1)

    wb = openpyxl.load_workbook(wb_path)
    written = 0
    for edge in new_edges:
        sheet_name = edge.get("target_sheet")
        cell_ref = edge.get("target_cell")
        comment_text = edge.get("description", "")
        author = edge.get("author", "mfq-tool")

        if not sheet_name or not cell_ref or not comment_text:
            continue

        if sheet_name not in wb.sheetnames:
            print(f"警告：sheet '{sheet_name}' 不存在，跳过", file=sys.stderr)
            continue

        ws = wb[sheet_name]
        cell = ws[cell_ref]
        cell.comment = Comment(comment_text, author)
        written += 1

    wb.save(wb_path)
    wb.close()
    print(f"已写入 {written} 条新耦合点到 {wb_path}")


def query_graph(graph_path: str, feature_name: str):
    """从序列化的图模型中查询某特性的所有耦合关系。"""
    try:
        # 支持 YAML 和 JSON
        if graph_path.endswith(".yaml") or graph_path.endswith(".yml"):
            try:
                import yaml
                with open(graph_path, "r", encoding="utf-8") as f:
                    graph = yaml.safe_load(f)
            except ImportError:
                print("警告：yaml 库不可用，尝试 JSON 格式", file=sys.stderr)
                with open(graph_path, "r", encoding="utf-8") as f:
                    graph = json.load(f)
        else:
            with open(graph_path, "r", encoding="utf-8") as f:
                graph = json.load(f)
    except Exception as e:
        print(f"错误：无法读取图模型文件 - {e}", file=sys.stderr)
        sys.exit(1)

    feature_lower = feature_name.lower()
    related_edges = []
    for edge in graph.get("edges", []):
        desc = edge.get("description", "").lower()
        if feature_lower in desc:
            related_edges.append(edge)

    print(f"特性 '{feature_name}' 相关耦合点：{len(related_edges)} 条")
    for i, e in enumerate(related_edges, 1):
        print(f"  [{i}] {e.get('cell', '?')} | {e.get('strength', '?')} | {e.get('description', '')[:80]}")

    return related_edges


def save_graph(graph: dict, output_path: str):
    """将图模型序列化保存。"""
    ext = Path(output_path).suffix.lower()
    if ext in (".yaml", ".yml"):
        try:
            import yaml
            with open(output_path, "w", encoding="utf-8") as f:
                yaml.dump(graph, f, allow_unicode=True, default_flow_style=False)
            print(f"图模型已保存为 YAML：{output_path}")
            return
        except ImportError:
            print("yaml 库不可用，回退为 JSON 格式", file=sys.stderr)
            output_path = str(Path(output_path).with_suffix(".json"))

    with open(output_path, "w", encoding="utf-8") as f:
        json.dump(graph, f, ensure_ascii=False, indent=2)
    print(f"图模型已保存为 JSON：{output_path}")


def main():
    parser = argparse.ArgumentParser(description="Excel 耦合矩阵读写工具")
    subparsers = parser.add_subparsers(dest="command", required=True)

    # read 子命令
    read_parser = subparsers.add_parser("read", help="读取 Excel 耦合矩阵")
    read_parser.add_argument("excel_path", help="Excel 文件路径")
    read_parser.add_argument("--output", "-o", default=None, help="输出图模型文件路径（默认 .mfq-work/f-analysis/coupling-graph.json）")

    # write 子命令
    write_parser = subparsers.add_parser("write", help="将新耦合点写回 Excel")
    write_parser.add_argument("excel_path", help="Excel 文件路径")
    write_parser.add_argument("--source", required=True, help="新耦合点数据文件路径（YAML/JSON）")

    # query 子命令
    query_parser = subparsers.add_parser("query", help="查询某特性的耦合关系")
    query_parser.add_argument("graph_path", help="图模型文件路径")
    query_parser.add_argument("--feature", required=True, help="特性名称")

    args = parser.parse_args()

    if args.command == "read":
        graph = read_coupling_matrix(args.excel_path)
        output = args.output or ".mfq-work/f-analysis/coupling-graph.json"
        os.makedirs(os.path.dirname(output) or ".", exist_ok=True)
        save_graph(graph, output)

    elif args.command == "write":
        source_path = args.source
        try:
            if source_path.endswith((".yaml", ".yml")):
                import yaml
                with open(source_path, "r", encoding="utf-8") as f:
                    new_edges = yaml.safe_load(f).get("edges", [])
            else:
                with open(source_path, "r", encoding="utf-8") as f:
                    new_edges = json.load(f).get("edges", [])
        except Exception as e:
            print(f"错误：无法读取源文件 - {e}", file=sys.stderr)
            sys.exit(1)
        write_coupling_to_excel(args.excel_path, new_edges)

    elif args.command == "query":
        query_graph(args.graph_path, args.feature)


if __name__ == "__main__":
    main()
